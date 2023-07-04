import math
import openpyxl


class LoginScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c
    def mulwertyu(self, num1, num2):
        result = math.ceil(num1 * num2)
        return result

    def read_excel_local_sheet_TestData(self, path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)
