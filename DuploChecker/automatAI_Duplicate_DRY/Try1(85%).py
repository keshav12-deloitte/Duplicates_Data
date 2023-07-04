import openai

openai.api_key = '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'

response = openai.ChatCompletion.create(
    model="gpt-3.5-turbo",
    messages=[
        {"role": "user", "content": """As a ML/Al model i am training you what exactly duplicate code in the code base is :
Pre Analysis: Go through all the classes which are provided and then 
The word Duplicate methods means :
1.If function names are same you should say those are duplicate methods
2.If the function implementation are same those are duplicate methods
3.If the function implementation is almost same but the variable name are different ,then also those are duplicate methods
4.If the logic of the code in different methods is almost same then also those are duplicate methods
5.If methods are achieving the same work then also those are duplicate methods
6.If methods are similar then also those are duplicate methods
7.If code logic and functionality are similar then also those are duplicate methods

Post Analysis: After finding the duplicates method, print the duplicate methods lines along with class name


"""},
        {"role": "user", "content": """  Below are the classes 
        
import pandas
import math
import pandas as pd
class HomeScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def subtraction_num(self, a, b):
        c = a - b
        return c

    def csv_reader(self):
        # reading the CSV file
        csvFile = pandas.read_csv('Giants.csv')
        # displaying the contents of the CSV file
        print(csvFile)

    def read_excelSheet_local(self, path):
        dataframe1 = pd.read_excel(path)
        print(dataframe1)


import math
import openpyxl

class LoginScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def subtraction_two_num(self, a, b):
        c = a - b
        return c

    def mul(self, num1, num2):
        result = math.ceil(num1 * num2)
        return result

    def qweqrqerq(self):
        num1 = int(input("Enter the first number: "))
        num2 = int(input("Enter the second number: "))
        # Sum of the two input numbers
        spasm = num1 + num2
        # Output the result
        print("The sum of", num1, "and", num2, "is", sum)
        return spasm

    def read_excel_local_sheet_TestData(self, path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)



import math


class paymentScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def readAndWriteText(self):
        file1 = open("myfile.txt", "w")
        L = ["This is Delhi \n", "This is Paris \n", "This is London \n"]
        file1.writelines(L)
        file1.close()
        # Append-adds at last
        file1 = open("myfile.txt", "a")  # append mode
        file1.write("Today \n")
        file1.close()
        file1 = open("myfile.txt", "r")
        print("Output of Readlines after appending")
        print(file1.readlines())
        print()
        file1.close()

    def mul_asfsaf(self, a, b):
        soo = math.ceil(a * b)
        return soo

    def rajajaijdaisjf(self):
        num1 = int(input("Enter the first number: "))
        num2 = int(input("Enter the second number: "))
        # Sum of the two input numbers
        sum = num1 + num2
        # Output the result
        print("The sum of", num1, "and", num2, "is", sum)
         
         find in the classes duplicate code, check for DRY and find code with same logic"""},

    ]
)

print(response['choices'][0]['message']['content'])

####################################### RESPONSE - 1  ##################################################################

# /Users/amananand7/Documents/automatAI_Duplicate_DRY/bin/python /Users/amananand7/PycharmProjects/automatAI_Duplicate_DRY/Try1(85%).py
# There are several duplicate methods within the provided classes which violate the DRY principle:
#
# 1. The `sum_two_numbers()` method is present in all three classes.
# 2. The `subtraction_num()` and `subtraction_two_num()` are similar methods in `HomeScreen` and `LoginScreen`, respectively.
# 3. The `mul()` method in `LoginScreen` and `mul_asfsaf()` in `paymentScreen` have a similar implementation.
# 4. The `read_excelSheet_local()` in `HomeScreen` and `read_excel_local_sheet_TestData()` in `LoginScreen` are similar methods.
# 5. The `qweqrqerq()` in `LoginScreen` and `rajajaijdaisjf()` in `paymentScreen` have a similar implementation.
#
# To avoid duplicating the code, we can create a single class with these methods and use inheritance to extend it in each of the three classes. We can also create separate utility functions to handle file reading and writing operations as well as math calculations, instead of repeating those operations in each class. This way, we can ensure that the code is modular and follows the DRY principle.
#
# Process finished with exit code 0




##################################### RESPONSE -2 ######################################################################

# /Users/amananand7/Documents/automatAI_Duplicate_DRY/bin/python /Users/amananand7/PycharmProjects/automatAI_Duplicate_DRY/Try1(85%).py
# In the given classes, we can find some instances of duplicate code.
#
# 1. `sum_two_numbers()` method is present in `HomeScreen`, `LoginScreen`, and `paymentScreen` classes. All these methods have the same functionality. Hence, we can consider them duplicate methods.
#
# 2. `subtraction_num()` method in `HomeScreen` and `subtraction_two_num()` in `LoginScreen` are similar in functionality. They can be considered duplicate methods.
#
# 3. `mul()` method in `LoginScreen` and `mul_asfsaf()` in `paymentScreen` are similar in functionality. They can be considered duplicate methods.
#
# 4. `qweqrqerq()` method in `LoginScreen` and `rajajaijdaisjf()` in `paymentScreen` are almost similar in functionality, with only minor differences like variable names. Hence, we can consider them duplicate methods.
#
# 5. `read_excelSheet_local()` in `HomeScreen` and `read_excel_local_sheet_TestData()` in `LoginScreen` have the same functionality of reading an excel sheet. Hence, they can be considered duplicate methods.
#
# To remove these duplications, we can create a common parent class and move the duplicate methods to it.
#
# Process finished with exit code 0




##################################### RESPONSE - 3 ######################################################################

