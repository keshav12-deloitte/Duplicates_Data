import math
import openpyxl


class LoginScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def subtraction_two_num(self, a, b):
        c = a - b
        return c

    def mulwertyu(self, num1, num2):
        result = math.ceil(num1 * num2)
        return result

    def qweqrqerq(self):
        num1 = int(input("Enter the first number: "))
        num2 = int(input("Enter the second number: "))
        # Sum of the two input numbers
        spasm = num1 + num2
        # Output the result
        print("The sum of", num1, "and", num2, "is", sum)
        return spasm

    def read_excel_local_sheet_TestData(self, path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)
