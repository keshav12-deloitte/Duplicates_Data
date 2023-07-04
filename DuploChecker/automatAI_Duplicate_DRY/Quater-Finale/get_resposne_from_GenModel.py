import openai
import openpyxl

# Constants
""" All the constants required """

# Set up your OpenAI API credentials
openai.api_key = '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'
# which model to use
model_name = "text-davinci-003"
fine_tune_model_name = "davinci:ft-deloitteus-2023-06-22-14-26-18"
model_name1 = "gpt-3.5-turbo-0301"
# excel_path
excel_path = "/Users/amananand7/PycharmProjects/automatAI_Duplicate_DRY/resources/Duplicate_generator_Test_Data.xlsx"
# sheet name
sheet_name = "selenium"


def read_write_excel(file_path, sheet_name, row_name, column_name, data=None):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    column_index = openpyxl.utils.column_index_from_string(column_name)
    if data is None:
        cell_value = sheet.cell(row=row_name, column=column_index).value
        workbook.close()
        return cell_value
    sheet.cell(row=row_name, column=column_index).value = data
    workbook.save(file_path)
    workbook.close()


new_prompt = """Please review the given INPUT and  check for duplicates, plagiarism, and code similarity. Additionally, please provide the output in the following format:
Method Name: Write the method name 
ClassesName: Write the classes name in which it is present
Confidence Level: Confidence level in percentage
Reason: Provide a brief explanation of why the code is considered a duplicate or code similarity 
Method Description : Write a concise description of the method
Method : Print the duplicated code or code logic similar here

Instruction : Give the above output only for those methods which are duplicates, plagiarism, and code similarity.
INPUT  : """
for i in range(2, 5):
    # input_code = read_write_excel()
    conversation_list = [
        {"role": "system", "content": "You are a coder"},
        {"role": "user", "content": """Please review the given INPUT and  check for duplicates, plagiarism, and code similarity. Additionally, please provide the output in the following format:
Method Name: Write the method name 
ClassesName: Write the classes name in which it is present
Confidence Level: Confidence level in percentage
Reason: Provide a brief explanation of why the code is considered a duplicate or code similarity 
Method Description : Write a concise description of the method
Method : Print the duplicated code or code logic similar here

Instruction : Give the above output only for those methods which are duplicates, plagiarism, and code similarity.
INPUT  :  "class BasePage:
        def __init__(self, driver: Union[Chrome, Firefox, Edge], wait: WebDriverWait):
            self.driver = driver
            self.wait = wait
    
        def edit_cookie(self, cookie_key: str, cookie_value: str):
            cookie = self.wait.until(
            lambda d: d.get_cookie(cookie_key),
            message=f"Cookie '{cookie_key}' does not exist within the given timeout")
            self.driver.delete_cookie(cookie_key)
            self.driver.add_cookie({
                "name": cookie['name'],
                "value": cookie_value,
                "domain": cookie['domain'],
                "path": cookie['path'],
                "secure": cookie['secure'],
                "expiry":(cookie['expiry'])
            })
    
        def click(self, locator: Tuple[By, str]) -> None:
            el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
            self._highlight_element(el, "green")
            el.click()
    
        def fill_text(self, locator: Tuple[By, str], txt: str) -> None:
            el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
            el.clear()
            self._highlight_element(el, "green")
            el.send_keys(txt)
    
        def clear_text(self, locator: Tuple[By, str]) -> None:
            el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
            el.clear()
    
        def scroll_to_bottom(self) -> None:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    
        def submit(self,  webelement: WebElement) -> None:
            self._highlight_element(webelement, "green")
            webelement.submit()
    
        def get_text(self, locator: Tuple[By, str]) -> str:
            el: WebElement = self.wait.until(expected_conditions.visibility_of_element_located(locator))
            self._highlight_element(el, "green")
            return el.text
    
        def whether_elem_displayed(self, webelement: WebElement) -> bool:
            try:
                return webelement.is_displayed()
            except StaleElementReferenceException:
                return False
            except NoSuchElementException:
                return False
    class Raniji:
        
    
        def __init__(self, driver: Union[Chrome, Firefox, Edge], wait: WebDriverWait):
            self.driver = driver
            self.wait = wait
    
        
        def click_webelement(self, locator: Tuple[By, str]) -> None:
            el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
            self._highlight_element(el, ""green"")
            el.click()
    
        def fill_text(self, locator: Tuple[By, str], txt: str) -> None:
            el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
            el.clear()
            self._highlight_element(el, ""green"")
            el.send_keys(txt)
    
        def clear_input(self, locator: Tuple[By, str]) -> None:
            el.clear()
    
        def submit_button(self,  webelement: WebElement) -> None:
            webelement.submit()
    
        def get_text(self, locator: Tuple[By, str]) -> str:
            el: WebElement = self.wait.until(expected_conditions.visibility_of_element_located(locator))
            self._highlight_element(el, ""green"")
            return el.text
    
    
        def is_elem_displayed(self, webelement: WebElement) -> bool:
            try:
                return webelement.is_displayed()
            except NoSuchElementException:
                return False" """},
        {"role": "assistant", "content": """There is considerable code duplication and code similarity between the methods defined in the two classes - `BasePage` and `Raniji`.
    
    Method Name: `click`
    ClassName: `BasePage`, `Raniji`
    Confidence Level: 100%
    Reason: Both classes have a method named `click` which takes the same input parameters and performs the same operation on the web element.
    
    Method Description: Waits for an element to be clickable and clicks it.
    Method:
    ```
    def click(self, locator: Tuple[By, str]) -> None:
    el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
    self._highlight_element(el, "green")
    el.click()
    
    def click_webelement(self, locator: Tuple[By, str]) -> None:
    el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
    self._highlight_element(el, ""green"")
    el.click()
    ```
    
    Method Name: `fill_text`
    ClassName: `BasePage`, `Raniji`
    Confidence Level: 100%
    Reason: Both classes have a method named `fill_text` which takes the same input parameters and performs the same operation on the web element.
    
    Method Description: Waits for an element to be clickable, clears the input field(if any), fills the input field with given `txt`
    Method:
    ```
    def fill_text(self, locator: Tuple[By, str], txt: str) -> None:
    el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
    el.clear()
    self._highlight_element(el, "green")
    el.send_keys(txt)
    
    def fill_text(self, locator: Tuple[By, str], txt: str) -> None:
    el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
    el.clear()
    self._highlight_element(el, ""green"")
    el.send_keys(txt)
    ```
    
    Method Name: `clear_text` ,   `clear_input`
    ClassName: `BasePage`,  `Raniji`
    Confidence Level: 67%
    Reason: Both classes have a method named `clear_text` , `clear_input` are similiar as they takes the same input parameters and performs the same operation on the web element but in `clear_text`  we  are waiting for an element to be clickable.
    
    Method Description: Waits for an element to be clickable, clears the input field(if any)
    Method:
    ```
    def clear_text(self, locator: Tuple[By, str]) -> None:
    el: WebElement = self.wait.until(expected_conditions.element_to_be_clickable(locator))
    el.clear()
    ```
    Method Description: Clears the input field(if any)
    Method:
    ```
    def clear_input(self, locator: Tuple[By, str]) -> None:
    el.clear()
    ```
    
    Method Name: `submit`
    ClassName: `BasePage`, `Raniji`
    Confidence Level: 100%
    Reason: Both classes have a method named `submit` which takes the same input parameters and performs the same operation on the web element.
    
    Method Description: Submits the web element.
    Method:
    ```
    def submit(self, webelement: WebElement) -> None:
    self._highlight_element(webelement, "green")
    webelement.submit()
    
    def submit_button(self, webelement: WebElement) -> None:
    webelement.submit()
    ```
    
    Method Name: `get_text`
    ClassName: `BasePage`, `Raniji`
    Confidence Level: 100%
    Reason: Both classes have a method named `get_text` which takes the same input parameters and performs the same operation on the web element.
    
    Method Description: Waits for an element to be visible and returns its text.
    Method:
    ```
    def get_text(self, locator: Tuple[By, str]) -> str:
    el: WebElement = self.wait.until(expected_conditions.visibility_of_element_located(locator))
    self._highlight_element(el, "green")
    return el.text
    
    def get_text(self, locator: Tuple[By, str]) -> str:
    el: WebElement = self.wait.until(expected_conditions.visibility_of_element_located(locator))
    self._highlight_element(el, ""green"")
    return el.text
    ```
    
    Method Name: `whether_elem_displayed`, `is_elem_displayed`
    ClassName: `BasePage`, `Raniji`
    Confidence Level: 71%
    Reason: Both classes have a method named `whether_elem_displayed` , `is_elem_displayed` are similiar as they takes the same input parameters and performs the same operation on the web element but in `whether_elem_displayed` is handling StaleElementReferenceException also.
    
    Method Description: Waits for an element to be visible and returns boolean value True or False.
    Method:
    ```
    def whether_elem_displayed(self, webelement: WebElement) -> bool:
    try:
    return webelement.is_displayed()
    except StaleElementReferenceException:
    return False
    except NoSuchElementException:
    return False
    ```
    
    Method Description: Waits for an element to be visible and returns boolean value True or False.
    Method:
    ```
    def is_elem_displayed(self, webelement: WebElement) -> bool:
    try:
    return webelement.is_displayed()
    except NoSuchElementException:
    return False
    ```"""},
        {"role":"user", "content": """Please review the given INPUT and check for duplicates, plagiarism, and code similarity. Additionally, please provide the output in the following format:
Method Name: Write the method name
ClassesName: Write the classes name in which it is present
Confidence Level: Confidence level in percentage
Reason: Provide a brief explanation of why the code is considered a duplicate or code similarity
Method Description : Write a concise description of the method
Method : Print the duplicated code or code logic similar here

Instruction : Give the above output only for those methods which are duplicates, plagiarism, and code similarity.
INPUT :  "from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class BasePageElement:

def find_element(self, locator):
return self.wait.until(EC.presence_of_element_located(locator))

def find_multiple_elements(self, locator):
return self.wait.until(EC.presence_of_all_elements_located(locator))

def first_name_empty:
if firstName.isEmpty():
showError("First name is required")


"from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class PageElement:

def find_element_by_Locator(self, locator):
return self.wait.until(EC.presence_of_element_located(locator))

def find_multiple_elements(self, locator):
return self.wait.until(EC.presence_of_all_elements_located(locator))

def last_name_empty:
if lastName.isEmpty():
showError("Last name is required") " """},
        {"role":"assistant", "content":"""There is code duplication and code similarity between the methods defined in the two classes `BasePageElement` and `PageElement`.

Method Name: `find_multiple_elements`
ClassName: `BasePageElement`, `PageElement`
Confidence Level: 100%
Reason: Both classes have a method named `find_multiple_elements` which takes the same input parameters and performs the same operation to locate multiple web elements.
Method Description: Waits for multiple elements to be present and returns the list of elements that match the locator.
Method:
```
def find_multiple_elements(self, locator):
return self.wait.until(EC.presence_of_all_elements_located(locator))
```

Method Name: `find_element_by_Locator`, `find_element`
ClassName: `BasePageElement`, `PageElement`
Confidence Level: 86%
Reason: Both classes have a method named `find_element_by_Locator` and `find_element` respectively which are similar, as they perform the operation of waiting for the web element to be present with the given locator.

Method Description: Waits for a single element to be present and returns the element that matches the locator.
Method:
```
def find_element(self, locator):
return self.wait.until(EC.presence_of_element_located(locator))

def find_element_by_Locator(self, locator):
return self.wait.until(EC.presence_of_element_located(locator))
```

Method Name: `first_name_empty`, `last_name_empty`
ClassName: `BasePageElement`, `PageElement`
Confidence Level: 100%
Reason: Both classes have a method named `first_name_empty`, `last_name_empty` respectively which are completely identical as they perform the operation of checking if the first name or the last name is empty.

Method Description: Checks if the first name or last name is empty and shows an error message if it is empty.
Method:
```
def first_name_empty():
if firstName.isEmpty():
showError("First name is required")


def last_name_empty():
if lastName.isEmpty():
showError("Last name is required")
```"""},
        {"role": "user",
         "content": new_prompt + '"' + read_write_excel(excel_path, sheet_name, row_name=i, column_name="B") + '"'}
    ]

    # answer1 = openai.Completion.create(
    #     model=model_name,
    #     prompt=new_prompt,
    # )

    # print(answer1['choices'][0]['text'])
    # print(new_prompt + '"' +read_write_excel(excel_path, sheet_name, row_name=i, column_name="C") + '"')

    # Generate a response from GPT-3.5 Turbo
    response = openai.ChatCompletion.create(
        model=model_name1,
        messages=conversation_list,
        max_tokens=888
    )

    # Get the assistant's reply
    reply = response.choices[0].message['content']
    read_write_excel(excel_path, sheet_name, row_name=i, column_name="D", data=reply)
    # Print the assistant's reply

    print("#################################################Rajaji##########################################")
    print(reply)
