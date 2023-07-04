import openai

# Set up your OpenAI API credentials
openai.api_key = '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'

# job_id = "ft-kPAvL8LtPwNFlGCeTQ9rzwtc"
# status = openai.FineTune.retrieve(id=job_id)["status"]
# print(status)

# result = openai.FineTune.list()
# print(f'Found {len(result.data)} finetune jobs.')
#
# print("Rajaji")
# print(result.__dict__)

# fine_tuned_model_id = "davinci:ft-deloitteus-2023-06-21-16-53-53"
# fine_tuned_model_id = result.data[0]['fine_tuned_model_id']


# fine_tuned_model_id = ModelId(fine_tuned_model_id)
# fine_tuned_model = openai.Model.get(fine_tuned_model_id)
# fine_tuned_model = openai.FineTune.get()
# print(fine_tuned_model)

# fine_tuned_model = openai.Model.get(fine_tuned_model_id)
# print(fine_tuned_model)

# fine_tuned_model = result.fine_tuned_model
# print(fine_tuned_model)
# Use the fine tune model to give output
new_prompt2 = """ "Find the duplicate code in the below python code:
import pandas
import math
import pandas as pd
class maths:

    def Addition_of_numbers(self, a, b):
        c = a + b
        return c

    def read_excelSheet_local(self, path):
        dataframe1 = pd.read_excel(path)
        print(dataframe1)


import math
import openpyxl

class LoginFunctionality:

    def Addition_of_numbers(self, a, b):
        c = a + b
        return c

    def mul(self, num1, num2):
        result = math.ceil(num1 * num2)
        return result

    def read_excel_local_sheet_TestData(self, path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)



import math


class PaymentScreen:

    def Addition_of_numbers(self, a, b):
        c = a + b
        return c

    def mul_asfsaf(self, a, b):
        soo = math.ceil(a * b)
        return soo
"""
new_prompt1 = """ Find the duplicate code in the below python code
import pandas
import math
import pandas as pd
class HomeScreen:

def sum_two_numbers(self, a, b):
c = a + b
return c

def subtraction_num(self, a, b):
c = a - b
return c

def csv_reader(self):
# reading the CSV file
csvFile = pandas.read_csv('Giants.csv')
# displaying the contents of the CSV file
print(csvFile)

def read_excelSheet_local(self, path):
dataframe1 = pd.read_excel(path)
print(dataframe1)


import math
import openpyxl

class LoginScreen:

def sum_two_numbers(self, a, b):
c = a + b
return c

def subtraction_two_num(self, a, b):
c = a - b
return c

def mul(self, num1, num2):
result = math.ceil(num1 * num2)
return result

def qweqrqerq(self):
num1 = int(input("Enter the first number: "))
num2 = int(input("Enter the second number: "))
# Sum of the two input numbers
spasm = num1 + num2
# Output the result
print("The sum of", num1, "and", num2, "is", sum)
return spasm

def read_excel_local_sheet_TestData(self, path):
dataframe = openpyxl.load_workbook(path)
dataframe1 = dataframe.active
for row in range(0, dataframe1.max_row):
for col in dataframe1.iter_cols(1, dataframe1.max_column):
print(col[row].value)



import math


class paymentScreen:

def sum_two_numbers(self, a, b):
c = a + b
return c

def readAndWriteText(self):
file1 = open("myfile.txt", "w")
L = ["This is Delhi \n", "This is Paris \n", "This is London \n"]
file1.writelines(L)
file1.close()
# Append-adds at last
file1 = open("myfile.txt", "a") # append mode
file1.write("Today \n")
file1.close()
file1 = open("myfile.txt", "r")
print("Output of Readlines after appending")
print(file1.readlines())
print()
file1.close()

def mul_asfsaf(self, a, b):
soo = math.ceil(a * b)
return soo

def rajajaijdaisjf(self):
num1 = int(input("Enter the first number: "))
num2 = int(input("Enter the second number: "))
# Sum of the two input numbers
sum = num1 + num2
# Output the result
print("The sum of", num1, "and", num2, "is", sum)"""

new_prompt = """Find the duplicate code in the below python code:
import pandas
import math
import pandas as pd
class Mathematics:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def subtraction_num(self, a, b):
        c = a - b
        return c

    def csv_reader(self):
        # reading the CSV file
        csvFile = pandas.read_csv('Giants.csv')
        # displaying the contents of the CSV file
        print(csvFile)

    def read_excelSheet_local(self, path):
        dataframe1 = pd.read_excel(path)
        print(dataframe1)


import math
import openpyxl

class LoginScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def subtraction_two_num(self, a, b):
        c = a - b
        return c

    def mul(self, num1, num2):
        result = math.ceil(num1 * num2)
        return result

    def qweqrqerq(self):
        num1 = int(input("Enter the first number: "))
        num2 = int(input("Enter the second number: "))
        # Sum of the two input numbers
        spasm = num1 + num2
        # Output the result
        print("The sum of", num1, "and", num2, "is", sum)
        return spasm

    def read_excel_local_sheet_TestData(self, path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                print(col[row].value)



import math


class PaymentScreen:

    def sum_two_numbers(self, a, b):
        c = a + b
        return c

    def readAndWriteText(self):
        file1 = open("myfile.txt", "w")
        L = ["This is Delhi", "This is Paris", "This is London "]
        file1.writelines(L)
        file1.close()
        # Append-adds at last
        file1 = open("myfile.txt", "a")  # append mode
        file1.write("Today")
        file1.close()
        file1 = open("myfile.txt", "r")
        print("Output of Readlines after appending")
        print(file1.readlines())
        print()
        file1.close()

    def mul_asfsaf(self, a, b):
        soo = math.ceil(a * b)
        return soo
"""
answer = openai.Completion.create(
    model="davinci:ft-deloitteus-2023-06-22-14-26-18",
    prompt=new_prompt,
    temperature=0,
    max_tokens=777,
    top_p=1,
    frequency_penalty=0,
    presence_penalty=0,
    stop=["END"]
)

print(answer['choices'][0]['text'])


# print(answer.__dict__)